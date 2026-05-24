"""解析 Bug 表格，并转换成统一的 JSONL 记录。"""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BUG_DIR = PROJECT_ROOT / "data" / "bugs"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "processed" / "bug_items.jsonl"


@dataclass
class BugItem:
    item_type: str
    source_file: str
    title: str
    bug_id: str | None = None
    bug_type: str | None = None
    status: str | None = None
    priority: str | None = None
    severity: str | None = None
    version: str | None = None
    description: str | None = None
    sheet: str | None = None
    row_index: int | None = None
    raw: dict[str, Any] | None = None


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def display_path(path: Path) -> str:
    """生成数据里优先使用项目相对路径。"""
    try:
        return str(path.resolve().relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path)


def normalize_header(value: str) -> str:
    text = value.replace(" ", "").replace("\n", "").strip().lower()
    mapping = {
        "问题id": "bug_id",
        "bugid": "bug_id",
        "id": "bug_id",
        "编号": "bug_id",
        "序号": "index",
        "问题名称": "title",
        "bug名称": "title",
        "标题": "title",
        "问题标题": "title",
        "类型": "bug_type",
        "问题类型": "bug_type",
        "状态": "status",
        "优先级": "priority",
        "严重级别": "severity",
        "级别": "severity",
        "版本": "version",
        "影响版本": "version",
        "定义": "definition",
        "详细描述": "detail",
        "bug描述": "description",
        "问题描述": "description",
        "描述": "description",
        "备注": "remark",
    }
    return mapping.get(text, text)


def find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[int, str]] | None:
    for row_index, row in enumerate(rows[:30]):
        headers = {index: normalize_header(clean_cell(cell)) for index, cell in enumerate(row) if clean_cell(cell)}
        values = set(headers.values())
        if "bug_id" in values and "title" in values:
            return row_index, headers
        if "severity" in values and ("definition" in values or "description" in values):
            return row_index, headers
    return None


def row_to_values(row: tuple[Any, ...], headers: dict[int, str]) -> dict[str, str]:
    return {field: clean_cell(row[index]) for index, field in headers.items() if index < len(row)}


def parse_bug_excel(path: Path) -> list[BugItem]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    source_file = display_path(path)
    items: list[BugItem] = []

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        header_info = find_header_row(rows)
        if not header_info:
            continue

        header_row_index, headers = header_info
        header_values = set(headers.values())
        is_severity_rule_sheet = "severity" in header_values and ("definition" in header_values or "description" in header_values)
        last_values: dict[str, str] = {}

        for row_index, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
            values = row_to_values(row, headers)
            if not any(values.values()):
                continue

            # read_only 模式下，合并单元格的值只会出现在第一行。
            for key in ("severity", "definition"):
                if not values.get(key) and last_values.get(key):
                    values[key] = last_values[key]
            for key, value in values.items():
                if value:
                    last_values[key] = value

            if is_severity_rule_sheet:
                severity = values.get("severity") or values.get("index")
                description = values.get("description") or values.get("detail") or values.get("definition")
                if not severity and not description:
                    continue
                items.append(
                    BugItem(
                        item_type="bug_severity_rule",
                        source_file=source_file,
                        title=f"{severity or '未分级'} 缺陷等级规则",
                        severity=severity or None,
                        description=description or None,
                        sheet=worksheet.title,
                        row_index=row_index,
                        raw=values,
                    )
                )
                continue

            title = values.get("title") or values.get("description") or values.get("bug_id") or ""
            if not title:
                continue

            items.append(
                BugItem(
                    item_type="bug_record",
                    source_file=source_file,
                    title=title,
                    bug_id=values.get("bug_id") or None,
                    bug_type=values.get("bug_type") or None,
                    status=values.get("status") or None,
                    priority=values.get("priority") or None,
                    severity=values.get("severity") or None,
                    version=values.get("version") or None,
                    description=values.get("description") or values.get("remark") or None,
                    sheet=worksheet.title,
                    row_index=row_index,
                    raw=values,
                )
            )

    workbook.close()
    return items


def parse_bug_file(path: Path) -> list[BugItem]:
    if path.suffix.lower() == ".xlsx":
        return parse_bug_excel(path)
    return []


def parse_all_bugs(bug_dir: Path = DEFAULT_BUG_DIR) -> list[BugItem]:
    items: list[BugItem] = []
    for path in sorted(bug_dir.iterdir()):
        if not path.is_file():
            continue
        parsed = parse_bug_file(path)
        print(f"{path.name}: {len(parsed)} items")
        items.extend(parsed)
    return items


def write_jsonl(items: list[BugItem], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as file:
        for item in items:
            file.write(json.dumps(asdict(item), ensure_ascii=False) + "\n")


def main() -> None:
    items = parse_all_bugs(DEFAULT_BUG_DIR)
    write_jsonl(items, DEFAULT_OUTPUT)
    counts: dict[str, int] = {}
    for item in items:
        counts[item.item_type] = counts.get(item.item_type, 0) + 1
    print(f"total: {len(items)} items")
    print(f"by_type: {counts}")
    print(f"output: {DEFAULT_OUTPUT}")


if __name__ == "__main__":
    main()
