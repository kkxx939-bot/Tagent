"""
测试用例文件处理。

目标：
1. Excel 用例文件：提取成结构化 test_case。
2. XMind 脑图文件：提取成测试点 test_point。
3. 统一输出 JSONL，后续才能做检索、RAG、用例生成参考。

先只处理 data/test_cases 下的 .xlsx 和 .xmind。
"""

from __future__ import annotations

import argparse
import json
import zipfile
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CASE_DIR = PROJECT_ROOT / "data" / "test_cases"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "processed" / "case_items.jsonl"


@dataclass
class CaseItem:
    """统一后的用例/测试点记录。"""

    item_type: str
    source_file: str
    title: str
    module: str | None = None
    case_id: str | None = None
    steps: str | None = None
    expected_result: str | None = None
    sheet: str | None = None
    path: list[str] | None = None
    raw: dict[str, Any] | None = None


def clean_cell(value: Any) -> str:
    """把 Excel 单元格值转成干净字符串。"""
    if value is None:
        return ""
    return str(value).strip()


def display_path(path: Path) -> str:
    """保存项目内相对路径，避免 JSONL 里出现本机绝对目录。"""
    try:
        return str(path.resolve().relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path)


def normalize_header(value: str) -> str:
    """把不同命名的表头归一成内部字段。"""
    text = value.replace(" ", "").replace("\n", "").strip().lower()
    mapping = {
        "用例编号": "case_id",
        "用例id": "case_id",
        "caseid": "case_id",
        "编号": "case_id",
        "测试项目": "module",
        "模块": "module",
        "所属模块": "module",
        "测试名称": "title",
        "用例标题": "title",
        "标题": "title",
        "测试点": "title",
        "操作步骤": "steps",
        "测试步骤": "steps",
        "步骤": "steps",
        "预期结果": "expected_result",
        "期望结果": "expected_result",
        "预期": "expected_result",
    }
    return mapping.get(text, text)


def find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[int, str]] | None:
    """
    查找 Excel 表头行。

    真实 Excel 经常前几行是标题、说明、空行，所以不能假设第一行就是表头。
    """
    required_like = {"title", "steps", "expected_result"}
    for row_index, row in enumerate(rows):
        headers = {index: normalize_header(clean_cell(cell)) for index, cell in enumerate(row) if clean_cell(cell)}
        values = set(headers.values())
        if "case_id" in values and (required_like & values):
            return row_index, headers
        if "module" in values and "title" in values:
            return row_index, headers
    return None


def parse_excel_cases(path: Path) -> list[CaseItem]:
    """解析 .xlsx 表格型测试用例。"""
    workbook = load_workbook(path, read_only=True, data_only=True)
    items: list[CaseItem] = []

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        header_info = find_header_row(rows[:30])
        if not header_info:
            continue

        header_row_index, headers = header_info
        last_module = ""

        for row in rows[header_row_index + 1 :]:
            values = {field: clean_cell(row[index]) for index, field in headers.items() if index < len(row)}
            if not any(values.values()):
                continue

            module = values.get("module") or last_module
            if module:
                last_module = module

            title = values.get("title") or values.get("case_id") or ""
            steps = values.get("steps", "")
            expected = values.get("expected_result", "")

            if not title and not steps and not expected:
                continue

            items.append(
                CaseItem(
                    item_type="test_case",
                    source_file=display_path(path),
                    sheet=worksheet.title,
                    case_id=values.get("case_id") or None,
                    module=module or None,
                    title=title,
                    steps=steps or None,
                    expected_result=expected or None,
                    raw=values,
                )
            )

    workbook.close()
    return items


def parse_xmind_cases(path: Path) -> list[CaseItem]:
    """解析 .xmind 脑图测试点。"""
    with zipfile.ZipFile(path) as archive:
        names = set(archive.namelist())
        if "content.json" in names:
            data = json.loads(archive.read("content.json").decode("utf-8"))
            return parse_xmind_json(path, data)
        if "content.xml" in names:
            root = ElementTree.fromstring(archive.read("content.xml"))
            return parse_xmind_xml(path, root)
    return []


def parse_xmind_json(path: Path, data: Any) -> list[CaseItem]:
    """解析新版 XMind 的 content.json。"""
    items: list[CaseItem] = []

    def walk(topic: dict[str, Any], parents: list[str]) -> None:
        title = clean_cell(topic.get("title"))
        current_path = parents + ([title] if title else [])
        if title and parents:
            items.append(
                CaseItem(
                    item_type="test_point",
                    source_file=display_path(path),
                    title=title,
                    module=parents[0] if parents else None,
                    path=current_path,
                )
            )

        children = topic.get("children", {})
        attached = children.get("attached", [])
        for child in attached:
            walk(child, current_path)

    sheets = data if isinstance(data, list) else [data]
    for sheet in sheets:
        root_topic = sheet.get("rootTopic") or sheet.get("topic")
        if root_topic:
            walk(root_topic, [])

    return items


def parse_xmind_xml(path: Path, root: ElementTree.Element) -> list[CaseItem]:
    """解析旧版 XMind 的 content.xml。"""
    items: list[CaseItem] = []

    def local_name(tag: str) -> str:
        return tag.split("}", 1)[-1]

    def topic_title(topic: ElementTree.Element) -> str:
        for child in topic:
            if local_name(child.tag) == "title":
                return clean_cell(child.text)
        return ""

    def direct_child_topics(topic: ElementTree.Element) -> list[ElementTree.Element]:
        result: list[ElementTree.Element] = []
        for children in topic:
            if local_name(children.tag) != "children":
                continue
            for topics in children:
                if local_name(topics.tag) != "topics":
                    continue
                for child in topics:
                    if local_name(child.tag) == "topic":
                        result.append(child)
        return result

    def walk(topic: ElementTree.Element, parents: list[str]) -> None:
        title = topic_title(topic)
        current_path = parents + ([title] if title else [])
        if title and parents:
            items.append(
                CaseItem(
                    item_type="test_point",
                    source_file=display_path(path),
                    title=title,
                    module=parents[0] if parents else None,
                    path=current_path,
                )
            )
        for child in direct_child_topics(topic):
            walk(child, current_path)

    for topic in root.iter():
        if local_name(topic.tag) == "topic":
            walk(topic, [])
            break

    return items


def parse_case_file(path: Path) -> list[CaseItem]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return parse_excel_cases(path)
    if suffix == ".xmind":
        return parse_xmind_cases(path)
    return []


def write_jsonl(items: list[CaseItem], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as file:
        for item in items:
            file.write(json.dumps(asdict(item), ensure_ascii=False) + "\n")


def parse_all_cases(case_dir: Path = DEFAULT_CASE_DIR) -> list[CaseItem]:
    items: list[CaseItem] = []
    for path in sorted(case_dir.iterdir()):
        if path.suffix.lower() not in {".xlsx", ".xmind"}:
            continue
        parsed = parse_case_file(path)
        print(f"{path.name}: {len(parsed)} items")
        items.extend(parsed)
    return items


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse test case Excel/XMind files into JSONL.")
    parser.add_argument("--case-dir", type=Path, default=DEFAULT_CASE_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    items = parse_all_cases(args.case_dir)
    write_jsonl(items, args.output)
    item_counts = Counter(item.item_type for item in items)
    print(f"total: {len(items)} items")
    print(f"by_type: {dict(item_counts)}")
    print(f"output: {args.output}")


if __name__ == "__main__":
    main()
