"""
需求文件处理。

目标：
1. 把 Word/PDF/Excel 等需求资料转成统一 JSONL。
2. Word 文档按章节抽取 requirement_section。
3. Excel 需求表按行抽取 requirement_table_row。
4. 暂时无法解析的文件输出 unsupported，方便后续补能力。

这一步只做“数据标准化”，不做 RAG、不做向量化。
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import tempfile
import zipfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_REQUIREMENTS_DIR = PROJECT_ROOT / "data" / "requirements"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "processed" / "requirement_items.jsonl"


@dataclass
class RequirementItem:
    """统一后的需求记录。"""

    item_type: str
    source_file: str
    title: str
    content: str | None = None
    section_title: str | None = None
    sheet: str | None = None
    row_index: int | None = None
    raw: dict[str, Any] | None = None
    error: str | None = None


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def display_path(path: Path) -> str:
    """保存项目内相对路径，避免 JSONL 里出现本机绝对目录。"""
    try:
        return str(path.resolve().relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path)


def is_heading(text: str, style: str = "") -> bool:
    """粗略判断一段文本是否像章节标题。"""
    if not text:
        return False
    normalized_style = style.lower()
    if "heading" in normalized_style or "title" in normalized_style:
        return True
    if len(text) > 80:
        return False
    patterns = [
        r"^第[一二三四五六七八九十\d]+[章节部分]",
        r"^\d+(\.\d+)*[、.．\s]",
        r"^[一二三四五六七八九十]+[、.．\s]",
    ]
    return any(re.match(pattern, text) for pattern in patterns)


def make_sections(source_file: str, document_title: str, paragraphs: list[tuple[str, str]]) -> list[RequirementItem]:
    """把段落合并为章节记录。"""
    items: list[RequirementItem] = []
    current_title = document_title
    buffer: list[str] = []

    def flush() -> None:
        content = clean_text("\n".join(buffer))
        if not content:
            return
        items.append(
            RequirementItem(
                item_type="requirement_section",
                source_file=source_file,
                title=current_title,
                section_title=current_title,
                content=content,
            )
        )

    for text, style in paragraphs:
        text = clean_text(text)
        if not text:
            continue
        if is_heading(text, style):
            flush()
            current_title = text
            buffer = []
        else:
            buffer.append(text)

    flush()
    return items


def parse_docx(path: Path) -> list[RequirementItem]:
    """解析 docx：抽段落和表格。"""
    source_file = display_path(path)
    document_title = path.stem
    paragraphs: list[tuple[str, str]] = []
    table_items: list[RequirementItem] = []

    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}

    with zipfile.ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read("word/document.xml"))

    body = root.find("w:body", ns)
    if body is None:
        return []

    for node in body:
        tag = node.tag.split("}", 1)[-1]
        if tag == "p":
            text = "".join(t.text or "" for t in node.findall(".//w:t", ns))
            style_node = node.find(".//w:pStyle", ns)
            style = style_node.attrib.get(f"{{{ns['w']}}}val", "") if style_node is not None else ""
            if clean_text(text):
                paragraphs.append((text, style))
        elif tag == "tbl":
            rows = []
            for tr in node.findall(".//w:tr", ns):
                cells = []
                for tc in tr.findall("./w:tc", ns):
                    cell_text = "".join(t.text or "" for t in tc.findall(".//w:t", ns))
                    cells.append(clean_text(cell_text))
                if any(cells):
                    rows.append(cells)

            if not rows:
                continue
            headers = rows[0]
            for index, row in enumerate(rows[1:], start=2):
                raw = {headers[i] if i < len(headers) and headers[i] else f"column_{i + 1}": row[i] for i in range(len(row))}
                table_items.append(
                    RequirementItem(
                        item_type="requirement_table_row",
                        source_file=source_file,
                        title=document_title,
                        content=" | ".join(cell for cell in row if cell),
                        row_index=index,
                        raw=raw,
                    )
                )

    return make_sections(source_file, document_title, paragraphs) + table_items


def parse_xlsx(path: Path) -> list[RequirementItem]:
    """解析 Excel 需求表：每一行输出一条 requirement_table_row。"""
    source_file = display_path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    items: list[RequirementItem] = []

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        header_index = None
        headers: list[str] = []

        for index, row in enumerate(rows[:30]):
            values = [clean_text(cell) for cell in row]
            if sum(bool(value) for value in values) >= 2:
                header_index = index
                headers = [value or f"column_{i + 1}" for i, value in enumerate(values)]
                break

        if header_index is None:
            continue

        for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            values = [clean_text(cell) for cell in row]
            if not any(values):
                continue
            raw = {headers[i] if i < len(headers) else f"column_{i + 1}": value for i, value in enumerate(values)}
            content = " | ".join(value for value in values if value)
            items.append(
                RequirementItem(
                    item_type="requirement_table_row",
                    source_file=source_file,
                    title=path.stem,
                    content=content,
                    sheet=worksheet.title,
                    row_index=row_index,
                    raw=raw,
                )
            )

    workbook.close()
    return items


def parse_text_file(path: Path) -> list[RequirementItem]:
    source_file = display_path(path)
    text = path.read_text(encoding="utf-8", errors="ignore")
    paragraphs = [(paragraph, "") for paragraph in re.split(r"\n\s*\n", text)]
    return make_sections(source_file, path.stem, paragraphs)


def parse_doc_with_textutil(path: Path) -> list[RequirementItem]:
    """用 macOS textutil 兜底解析旧 .doc 文件。"""
    source_file = display_path(path)
    with tempfile.TemporaryDirectory() as temp_dir:
        output_path = Path(temp_dir) / f"{path.stem}.txt"
        result = subprocess.run(
            ["textutil", "-convert", "txt", str(path), "-output", str(output_path)],
            capture_output=True,
            text=True,
            check=False,
        )
        if result.returncode != 0 or not output_path.exists():
            return [
                RequirementItem(
                    item_type="unsupported",
                    source_file=source_file,
                    title=path.stem,
                    error=clean_text(result.stderr) or "textutil conversion failed",
                )
            ]
        text = output_path.read_text(encoding="utf-8", errors="ignore")
    paragraphs = [(paragraph, "") for paragraph in re.split(r"\n\s*\n", text)]
    return make_sections(source_file, path.stem, paragraphs)


def parse_pdf(path: Path) -> list[RequirementItem]:
    """PDF 解析入口：当前环境没有 PDF 解析库时，输出 unsupported。"""
    source_file = display_path(path)
    try:
        from pypdf import PdfReader
    except ImportError:
        return [
            RequirementItem(
                item_type="unsupported",
                source_file=source_file,
                title=path.stem,
                error="PDF parsing requires pypdf or another PDF text extraction tool.",
            )
        ]

    reader = PdfReader(str(path))
    paragraphs: list[tuple[str, str]] = []
    for page_index, page in enumerate(reader.pages, start=1):
        text = clean_text(page.extract_text() or "")
        if text:
            paragraphs.append((f"第 {page_index} 页\n{text}", ""))
    return make_sections(source_file, path.stem, paragraphs)


def parse_requirement_file(path: Path) -> list[RequirementItem]:
    suffix = path.suffix.lower()
    if suffix == ".docx":
        return parse_docx(path)
    if suffix == ".xlsx":
        return parse_xlsx(path)
    if suffix == ".doc":
        return parse_doc_with_textutil(path)
    if suffix == ".pdf":
        return parse_pdf(path)
    if suffix in {".txt", ".md", ".csv", ".yaml", ".yml"}:
        return parse_text_file(path)
    return [
        RequirementItem(
            item_type="unsupported",
            source_file=display_path(path),
            title=path.stem,
            error=f"Unsupported requirement file type: {suffix}",
        )
    ]


def parse_all_requirements(requirements_dir: Path = DEFAULT_REQUIREMENTS_DIR) -> list[RequirementItem]:
    items: list[RequirementItem] = []
    for path in sorted(requirements_dir.iterdir()):
        if not path.is_file():
            continue
        parsed = parse_requirement_file(path)
        print(f"{path.name}: {len(parsed)} items")
        items.extend(parsed)
    return items


def write_jsonl(items: list[RequirementItem], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as file:
        for item in items:
            file.write(json.dumps(asdict(item), ensure_ascii=False) + "\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse requirement files into JSONL.")
    parser.add_argument("--requirements-dir", type=Path, default=DEFAULT_REQUIREMENTS_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    items = parse_all_requirements(args.requirements_dir)
    write_jsonl(items, args.output)
    counts: dict[str, int] = {}
    for item in items:
        counts[item.item_type] = counts.get(item.item_type, 0) + 1
    print(f"total: {len(items)} items")
    print(f"by_type: {counts}")
    print(f"output: {args.output}")


if __name__ == "__main__":
    main()
